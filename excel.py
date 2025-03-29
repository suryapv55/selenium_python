import time
import requests
from docutils.parsers.rst.directives.misc import Class

from selenium.webdriver.chrome import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import select
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
import openpyxl

#file="C:\\Users\\Sinergia\\Desktop\\DDTest.xlsx"

wb = openpyxl.load_workbook(file)
sheet = wb.active

row = sheet.max_row
col = sheet.max_column
print(row,col)
#Write Excel
sheet.cell(5,1).value='admin@sample.com'
sheet.cell(5,2).value='admin55'
wb.save(file)

#Read data from excel
for r in range(1, row + 1):
    for c in range(1, col + 1):
        print(f"{sheet.cell(row=r, column=c).value}", end='   ')
    print()

