import openpyxl
from openpyxl.styles import PatternFill


class ExcelUtils:
    @staticmethod
    def getRowCount(file, sheetName):
        wb = openpyxl.load_workbook(file)
        sheet = wb[sheetName]
        return sheet.max_row

    @staticmethod
    def getColCount(file, sheetName):
        wb = openpyxl.load_workbook(file)
        sheet = wb[sheetName]
        return sheet.max_column

    @staticmethod
    def readData(file, sheetName, rowno, colno):
        wb = openpyxl.load_workbook(file)
        sheet = wb[sheetName]
        return sheet.cell(row=rowno, column=colno).value

    @staticmethod
    def writeData(file, sheetName, rowno, colno, data):
        wb = openpyxl.load_workbook(file)
        sheet = wb[sheetName]
        sheet.cell(row=rowno, column=colno).value = data
        wb.save(file)

    @staticmethod
    def fillGreen(file, sheetName, rowno, colno):
        wb = openpyxl.load_workbook(file)
        sheet = wb[sheetName]
        green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        sheet.cell(row=rowno, column=colno).fill = green
        wb.save(file)

    @staticmethod
    def fillRed(file, sheetName, rowno, colno):
        wb = openpyxl.load_workbook(file)
        sheet = wb[sheetName]
        red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        sheet.cell(row=rowno, column=colno).fill = red
        wb.save(file)
